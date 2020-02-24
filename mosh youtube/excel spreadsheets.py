# python automation 

# use excel sheet 
# and work with excel sheet
# access row and column 
# for these we install package
#  >> pip install openpyxl

import openpyxl as xl
# make object of work book
wb = xl.load_workbook('spread_sheet.xlsx')
sheet =  wb['Sheet1']

# two way to access cell from sheet 
cell = sheet['a1']
cell = sheet.cell(1,1)
print(cell.value)
#sheet['d1']='aaam'
#cell = sheet['d1']
#print(cell.value)

# print all value from sheet
 
for i in range(1,sheet.max_row+1):
    # for j in range(1,sheet.max_col+1): 
    print(sheet.cell(i,1).value)

# update value in sheet and then save** file
 
# add new colum which contain
# which contain 90% of price colum
#print(sheet.max_col)
print()
for row in range(2,sheet.max_row+1):
    sheet.cell(row,4).value = sheet.cell(row,3).value*.9
    print(sheet.cell(row,4).value)

#wb.save('spread_sheet1.xlsx')

# draw a bar chart form data
# for that one we have to include some class from openpyxl.chart 
#  class = BarChart , Reference
from openpyxl.chart import BarChart, Reference 

# take data or refernce from sheet 
value_for_chart = Reference(sheet,min_row=2,
                             max_row=sheet.max_row, 
                             min_col=4,
                             max_col=4 )

# use BarChart class to draw 
# create object of BarChart class and add data
chart = BarChart()
chart.add_data(value_for_chart)
# add chart to sheet
sheet.add_chart(chart,'e2')
#save file(work book) 
wb.save('spread_sheet2.xlsx')





